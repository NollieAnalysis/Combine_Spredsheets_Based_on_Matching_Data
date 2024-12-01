'''
code for picklist meeting (SJ)
'''
# imports
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Alignment, Font
from copy import copy


# Function to copy formatting from one sheet to another
def copy_formatting(original_ws, new_ws):
    for row in original_ws.iter_rows():
        for cell in row:
            new_cell = new_ws[cell.coordinate]
            # Copy each individual style attribute
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format  # Direct assignment
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

# Set header rows dynamically
def set_header_row(file_path, sheet_names, key_column_first):
    header_rows = {}
    for sheet_name in sheet_names:
        df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        header_row = None
        for i, row in df_raw.iterrows():
            if key_column_first in row.values:
                header_row = i
                break
        header_rows[sheet_name] = header_row
    return header_rows

# Merge spreadsheets and preserve formatting
def combine_spreadsheets_preserve_formatting(first_spreadsheet_path, second_spreadsheet_path, output_spreadsheet_path, key_column_first, key_column_second, columns_to_copy, sheet_names):
    # Load the second spreadsheet
    df2 = pd.read_excel(second_spreadsheet_path)
    
    # Detect header rows in the first spreadsheet
    header_rows = set_header_row(first_spreadsheet_path, sheet_names, key_column_first)

    # Open the original workbook
    original_wb = load_workbook(first_spreadsheet_path)
    combined_wb = load_workbook(first_spreadsheet_path)  # Start with a copy of the original

    for sheet_name in sheet_names:
        # Get header row
        header_row = header_rows[sheet_name]

        # Load the specific sheet with the correct header row
        df1 = pd.read_excel(first_spreadsheet_path, sheet_name=sheet_name, header=header_row)

        # Store original row positions
        df1['Original Row Index'] = df1.index

        # Merge the sheets
        df_combined = pd.merge(
            df1,
            df2[[key_column_second] + columns_to_copy],
            how='left',
            left_on=key_column_first,
            right_on=key_column_second
        )

        # Drop redundant key column
        df_combined = df_combined.drop(columns=[key_column_second])

        # Restore original order
        df_combined = df_combined.sort_values('Original Row Index').drop(columns=['Original Row Index'])

        # Update the data in the combined workbook
        combined_ws = combined_wb[sheet_name]
        original_ws = original_wb[sheet_name]

        # Write headers explicitly
        for col_idx, header in enumerate(df_combined.columns, start=1):
            combined_ws.cell(row=header_row + 1, column=col_idx, value=header)

        # Write the data back while maintaining formatting
        for row_idx, row in enumerate(df_combined.itertuples(index=False), start=header_row + 2):
            for col_idx, value in enumerate(row, start=1):
                combined_ws.cell(row=row_idx, column=col_idx, value=value)

        # Copy formatting from original sheet to new sheet
        copy_formatting(original_ws, combined_ws)

    # Save the new workbook
    combined_wb.save(output_spreadsheet_path)
    print(f"Combined spreadsheet saved as {output_spreadsheet_path}")


# fill in file locations and combination criteria
user_first_spread_name = str(input("Enter name of SJ picklist spreadsheet file (including file type): "))
first_spreadsheet_path = f"/{user_first_spread_name}"

user_second_spread_name = str(input("Enter name of personal meeting prep spreadsheet file (including file type): "))
second_spreadsheet_path = f"/{user_second_spread_name}"

output_spreadsheet_path = f"/{user_first_spread_name}(ADDED_INFO).xlsx"

key_column_first = 'Order'  # The column to match in the first spreadsheet
key_column_second = 'Order Number'  # The column to match in the second spreadsheet
columns_to_copy = ['Permit Facilitator', 'Permit Facilitator LAN ID', 'Latest Expiration Date', 'E Permit Status', 'EPWC Comments']  # Columns from second spreadsheet to copy to first
sheet_names = ['CINNABAR', 'EDENVALE', 'GC']  # The name or index of the sheets you want to use in the first spreadsheet

combine_spreadsheets_preserve_formatting(first_spreadsheet_path, second_spreadsheet_path, output_spreadsheet_path, key_column_first, key_column_second, columns_to_copy, sheet_names)